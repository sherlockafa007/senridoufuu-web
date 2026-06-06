"""
文件对比分析工具 — 本地 Gradio 版
支持大型年报、招股书（PDF / Word / Excel / CSV）
使用分层摘要策略处理超长文件，无超时限制
"""

import os
import sys
import gradio as gr
from openai import OpenAI

import pdfplumber
import docx as python_docx
import openpyxl

# ── 配置 ──────────────────────────────────────────────────
QWEN_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
QWEN_MODEL    = "qwen-plus"

CHUNK_SIZE          = 8000   # 每段字符数（送给 Qwen 的文本块）
MAX_CHUNKS_PER_DOC  = 10     # 每份文件最多处理的段数（约 80K 字符 / ~300 页）
SUMMARY_MAX_TOKENS  = 700    # 每段摘要的最大输出 token
FINAL_MAX_TOKENS    = 3000   # 最终对比报告的最大输出 token

# ── 文字提取 ──────────────────────────────────────────────

def extract_text(filepath: str) -> str:
    ext = filepath.rsplit(".", 1)[-1].lower()
    if ext == "pdf":
        return _extract_pdf(filepath)
    elif ext == "docx":
        return _extract_word(filepath)
    elif ext in ("xlsx", "xls"):
        return _extract_excel(filepath)
    elif ext == "csv":
        with open(filepath, encoding="utf-8", errors="ignore") as f:
            return f.read()
    return ""


def _extract_pdf(filepath: str) -> str:
    pages = []
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                pages.append(t)
    return "\n".join(pages)


def _extract_word(filepath: str) -> str:
    doc = python_docx.Document(filepath)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_excel(filepath: str) -> str:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    parts = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 1000:
                rows.append("…（已截断，仅显示前 1000 行）")
                break
            cells = "\t".join("" if v is None else str(v) for v in row)
            if cells.strip():
                rows.append(cells)
        if rows:
            parts.append(f"[Sheet: {name}]\n" + "\n".join(rows))
    return "\n\n".join(parts)


# ── Qwen 调用 ─────────────────────────────────────────────

def _qwen(client: OpenAI, system: str, user: str, max_tokens: int = 800) -> str:
    resp = client.chat.completions.create(
        model=QWEN_MODEL,
        messages=[
            {"role": "system", "content": system},
            {"role": "user",   "content": user},
        ],
        max_tokens=max_tokens,
        temperature=0.3,
    )
    return resp.choices[0].message.content.strip()


# ── 分层摘要（处理超长文件的核心） ────────────────────────

EXTRACT_SYSTEM = (
    "你是专业的商业分析师。"
    "请从以下文件内容中提取核心商业信息，包括：主要业务、收入来源、"
    "关键财务/运营数据、战略重点、风险因素。"
    "输出要点格式，简洁、有数据支撑。"
)

MERGE_SYSTEM = (
    "以下是同一份文件的分段摘要。请将其整合为一份完整、有结构的商业摘要，"
    "保留所有关键数据和要点，去掉重复内容。"
)


def summarize_document(client: OpenAI, text: str, filename: str, log_fn=None) -> str:
    total_chars = len(text)

    if log_fn:
        log_fn(f"  {filename}：共 {total_chars:,} 字符")

    # 短文件：单次调用
    if total_chars <= CHUNK_SIZE * 2:
        return _qwen(client, EXTRACT_SYSTEM, f"文件：{filename}\n\n{text}", SUMMARY_MAX_TOKENS * 2)

    # 长文件：分段摘要
    chunks = [
        text[i : i + CHUNK_SIZE]
        for i in range(0, min(total_chars, CHUNK_SIZE * MAX_CHUNKS_PER_DOC), CHUNK_SIZE)
    ]
    skipped = total_chars - CHUNK_SIZE * MAX_CHUNKS_PER_DOC
    if skipped > 0 and log_fn:
        log_fn(f"  ⚠ {filename} 文件过大，仅分析前 {CHUNK_SIZE * MAX_CHUNKS_PER_DOC:,} 字符（约前 {MAX_CHUNKS_PER_DOC * 3} 页）")

    chunk_summaries = []
    for i, chunk in enumerate(chunks):
        if log_fn:
            log_fn(f"  正在分析第 {i + 1}/{len(chunks)} 段…")
        s = _qwen(
            client,
            EXTRACT_SYSTEM,
            f"《{filename}》第 {i + 1}/{len(chunks)} 段\n\n{chunk}",
            SUMMARY_MAX_TOKENS,
        )
        chunk_summaries.append(s)

    # 合并：段数少时直接拼接；段数多时再做一次汇总
    if len(chunk_summaries) <= 3:
        return "\n\n".join(chunk_summaries)

    if log_fn:
        log_fn(f"  正在汇总 {filename} 的各段摘要…")
    merged = _qwen(
        client,
        MERGE_SYSTEM,
        f"文件：{filename}\n\n" + "\n---\n".join(chunk_summaries),
        SUMMARY_MAX_TOKENS * 2,
    )
    return merged


# ── 主分析函数 ────────────────────────────────────────────

REPORT_SYSTEM = (
    "你是资深商业分析师，擅长多企业/多业务模式的横向对比。"
    "请基于以下各文件摘要，撰写一份深度对比分析报告。"
    "要求：结构清晰（用标题分层）、有具体数据支撑、"
    "指出各方核心异同、最终给出有价值的商业洞察和建议。"
    "用中文撰写。"
)


def analyze(api_key: str, files, question: str, progress=gr.Progress()) -> str:
    api_key = api_key.strip()
    if not api_key:
        return "❌ 请填写 Qwen API Key"
    if not files:
        return "❌ 请上传至少一个文件"

    client = OpenAI(api_key=api_key, base_url=QWEN_BASE_URL)
    logs   = []

    def log(msg: str):
        logs.append(msg)

    n = len(files)
    summaries = []

    for i, file in enumerate(files):
        fname = os.path.basename(file.name)
        progress((i * 2) / (n * 2 + 1), desc=f"提取文字：{fname}")
        log(f"\n📄 处理文件 {i + 1}/{n}：{fname}")

        try:
            text = extract_text(file.name)
        except Exception as e:
            log(f"  ❌ 提取失败：{e}")
            summaries.append(f"【{fname}】\n提取文字失败：{e}")
            continue

        if not text.strip():
            log(f"  ⚠ 未提取到文字（可能是扫描版 PDF，不含文字层）")
            summaries.append(f"【{fname}】\n（未能提取文字，可能是扫描版 PDF）")
            continue

        progress((i * 2 + 1) / (n * 2 + 1), desc=f"AI 分析：{fname}")

        try:
            summary = summarize_document(client, text, fname, log_fn=log)
            summaries.append(f"【{fname}】\n{summary}")
            log(f"  ✓ 完成")
        except Exception as e:
            log(f"  ❌ 分析失败：{e}")
            summaries.append(f"【{fname}】\n分析失败：{e}")

    if not any("【" in s and "失败" not in s for s in summaries):
        return "❌ 没有成功处理任何文件\n\n" + "\n".join(logs)

    progress(n * 2 / (n * 2 + 1), desc="生成综合对比报告…")
    log("\n📊 正在生成综合对比报告…")

    sep      = "\n\n" + "─" * 40 + "\n\n"
    combined = sep.join(summaries)
    q        = question.strip() or (
        "请对以上文件进行综合对比分析，指出各方的核心业务模式异同、"
        "关键指标对比和战略洞察。"
    )

    try:
        report = _qwen(
            client,
            REPORT_SYSTEM,
            f"各文件摘要如下：\n\n{combined}\n\n分析要求：{q}",
            FINAL_MAX_TOKENS,
        )
    except Exception as e:
        report = f"❌ 生成报告失败：{e}"

    progress(1.0, desc="完成")
    log("✓ 报告生成完毕")

    process_log = "\n".join(logs)
    return f"{report}\n\n\n{'─'*60}\n【处理日志】\n{process_log}"


# ── Gradio UI ─────────────────────────────────────────────

CSS = """
.output-text textarea { font-family: 'Noto Sans SC', sans-serif; line-height: 1.8; }
"""

with gr.Blocks(title="文件对比分析工具", css=CSS) as app:

    gr.Markdown(
        """# 文件对比分析工具
**本地运行，无超时限制。** 支持年报、招股书、研究报告等大型文件（PDF / Word / Excel / CSV）。

**原理**：本地提取文字 → 分段送 Qwen 提炼要点 → 汇总后生成对比报告。
扫描版 PDF（无文字层）无法自动提取，需先用 OCR 工具转换。"""
    )

    with gr.Row():
        api_key_box = gr.Textbox(
            label="Qwen API Key",
            placeholder="sk-xxxxxxxxxxxxxxxx",
            type="password",
            value=os.getenv("QWEN_API_KEY", ""),
            scale=1,
        )

    file_box = gr.File(
        file_count="multiple",
        label="上传文件（可同时上传多份，每份支持数百页）",
        file_types=[".pdf", ".docx", ".xlsx", ".xls", ".csv"],
    )

    question_box = gr.Textbox(
        label="分析问题（可留空，留空则综合对比）",
        placeholder="例：比较三条地铁的盈利模式、客流量趋势和未来扩张策略",
        lines=2,
    )

    submit_btn = gr.Button("开始分析", variant="primary", size="lg")

    output_box = gr.Textbox(
        label="分析报告",
        lines=35,
        show_copy_button=True,
        placeholder="报告将在此显示，包含分析结果和处理日志…",
        elem_classes=["output-text"],
    )

    submit_btn.click(
        fn=analyze,
        inputs=[api_key_box, file_box, question_box],
        outputs=output_box,
    )

if __name__ == "__main__":
    app.launch(share=False, inbrowser=True)
